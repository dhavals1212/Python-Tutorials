import openpyxl as xl
#This will open openpyxl as xl.

from openpyxl.chart import BarChart, Reference
# from the openpyxl was called chart option.
# from that chart option we import two classes BarChart and Reference.

def process_workbook(filename):
    # we define process_workbook as the function. It takes parameter of filename.

    wb = xl.load_workbook(filename)
    #xl.load_workbook will open the filename file and store it in wb.

    sheet = wb['Sheet1']
    #sheet variable will be storing the 'Sheet1' the name of the sheet in the variable. It is important to have 'wb'.

    #cell = sheet['a1']
    #This is to import a1 cell.
    #Now, cell = sheet.cell(1, 1) can do the same.

    for row in range(2, sheet.max_row + 1):
        #this will choose range of row 2 to row 4 for further process.
        # here sheet.max_row maintains that the maximum row the sheet has are 4.

        cell = sheet.cell(row, 3)
        #in the sheet, the cell key is called for the row's 3rd colum, which is stored in 'cell'.

        corrected_price = cell.value * 0.9
        #the formulae for the corrected_price which will store all 3 prices in the values.

        corrected_price_cell = sheet.cell(row,4)
        #this locates the cell values for the 4th row for newer values to be stored.

        corrected_price_cell.value = corrected_price
        #corrected prices for all 3 datas will be now stored in the corrected_price_cell.value.

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    #Reference class is called. It first requires the sheet.
    #Further it requires the minimum row number and maximum row number,
    #For it not to take other values except 4, we also define min and max column value.

    chart = BarChart()
    #Object 'chart' is created in which the BarChart() class is stored.

    chart.add_data(values)
    #Here the chart.add_data(values) adds the values referenced above to the chart as data.

    sheet.add_chart(chart, 'e2')
    #In sheet, the add_chart will ask what chart, and will get reference cell to add the chart to.

    wb.save('transactions2.xlsx')
    #The wb.save saves the workbook. Inside the key function, name of the file is changed.